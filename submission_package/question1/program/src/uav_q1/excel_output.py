from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "4472C4"
LIGHT = "D9EAF7"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E2F3")


def style_header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_workbook(results: dict, output_path: Path, config: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "问题一：多无人机协同巡检结果"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(bold=True, color=WHITE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["测试算例", "无人机数量 N", "最长工作时间 Tmax (h)", "最短工作时间 Tmin (h)"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    style_header(ws[3])
    for row, (case_name, case) in enumerate(results.items(), 4):
        ws.cell(row, 1, case_name)
        ws.cell(row, 2, case["N"])
        ws.cell(row, 3, case["Tmax"])
        ws.cell(row, 4, case["Tmin"])
        ws.cell(row, 3).number_format = ws.cell(row, 4).number_format = "0.0000"
    ws["A10"] = "最终计数规则"
    ws["B10"] = "到达并作业 5 min 计 1 次；原地停留不重复计数；离开后返回可再次计数；不同无人机可同时巡检。"
    ws.merge_cells("B10:D11")
    ws["A12"] = "时间公式"
    ws["B12"] = "T = Flight Distance / 55 + Inspection Count × 5 / 60"
    ws.merge_cells("B12:D12")
    for col, width in zip("ABCD", [20, 25, 30, 30]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    for case_name, case in results.items():
        sh = wb.create_sheet(case_name)
        sh.sheet_view.showGridLines = False
        max_len = max(len(x["sequence"]) for x in case["routes"])
        headers = ["UAV ID", *[f"{i}th Inspection Point" for i in range(1, max_len + 1)],
                   "Flight Distance (km)", "Inspection Count", "Working Time (h)"]
        for col, value in enumerate(headers, 1):
            sh.cell(5, col, value)
        style_header(sh[5])
        sh["A1"] = f"{case_name} 详细调度方案（0 表示基地）"
        end_col = get_column_letter(len(headers))
        sh.merge_cells(f"A1:{end_col}1")
        sh["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sh["A1"].font = Font(bold=True, color=WHITE, size=14)
        sh["A1"].alignment = Alignment(horizontal="center")
        sh["A2"] = "N"; sh["B2"] = case["N"]
        sh["C2"] = "Tmax (h)"; sh["D2"] = case["Tmax"]
        sh["E2"] = "Tmin (h)"; sh["F2"] = case["Tmin"]
        for row, route in enumerate(case["routes"], 6):
            values = [route["uav"], *route["sequence"], *([None] * (max_len - len(route["sequence"]))) ,
                      route["distance_km"], route["service_count"]]
            for col, value in enumerate(values, 1): sh.cell(row, col, value)
            distance_col = 2 + max_len
            count_col = distance_col + 1
            time_col = distance_col + 2
            sh.cell(row, time_col, f"={get_column_letter(distance_col)}{row}/55+{get_column_letter(count_col)}{row}*5/60")
            sh.cell(row, distance_col).number_format = "0.000"
            sh.cell(row, time_col).number_format = "0.000000"
        sh.freeze_panes = "B6"
        sh.column_dimensions["A"].width = 10
        for c in range(2, 2 + max_len): sh.column_dimensions[get_column_letter(c)].width = 11
        for c in range(2 + max_len, len(headers) + 1): sh.column_dimensions[get_column_letter(c)].width = 19
        sh.auto_filter.ref = f"A5:{end_col}{5 + len(case['routes'])}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
