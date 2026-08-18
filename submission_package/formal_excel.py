"""按赛题给定的五列示例模板生成正式路线工作簿。

正式结果只包含 UAV 编号和巡检点序列。基地编号 0、距离、时间、目标函数、
算法说明和逐事件记录均属于复核材料，不写入正式 result1/2/3.xlsx。
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


CASE_NAMES = ("Case1", "Case2", "Case3", "Case4")


def _case_payload(results: dict, case_name: str, problem: int) -> dict:
    case = results[case_name]
    if problem == 3:
        case = case["formal_fixed_N0"]
    return case


def _inspection_sequence(route: dict) -> list[int]:
    sequence = [int(x) for x in route["sequence"]]
    if len(sequence) < 3 or sequence[0] != 0 or sequence[-1] != 0:
        raise ValueError(f"UAV {route.get('uav')}: 路线必须以基地 0 开始并结束")
    visits = sequence[1:-1]
    if any(point == 0 for point in visits):
        raise ValueError(f"UAV {route.get('uav')}: 路线中间不能出现基地 0")
    return visits


def export_formal_workbook(
    results: dict,
    output_path: Path,
    template_path: Path,
    problem: int,
) -> None:
    """从已验证 JSON 生成仅含正式路线表的 resultX.xlsx。"""
    if tuple(results) != CASE_NAMES and set(results) != set(CASE_NAMES):
        raise ValueError(f"结果算例必须恰好为 {CASE_NAMES}")
    if problem not in (1, 2, 3):
        raise ValueError("problem 只能为 1、2 或 3")

    workbook = load_workbook(template_path)
    source = workbook.active
    sheets = [source]
    for _ in range(3):
        sheets.append(workbook.copy_worksheet(source))

    template_style = copy(source["A1"]._style)
    template_width_a = source.column_dimensions["A"].width or 7.0
    template_width_route = source.column_dimensions["B"].width or 18.0

    for worksheet, case_name in zip(sheets, CASE_NAMES):
        worksheet.title = case_name
        case = _case_payload(results, case_name, problem)
        routes = sorted(case["routes"], key=lambda item: int(item["uav"]))
        if len(routes) != int(case["N"]):
            raise ValueError(f"{case_name}: N 与路线数量不一致")

        visits_by_route = [_inspection_sequence(route) for route in routes]
        max_stops = max(len(visits) for visits in visits_by_route)

        # 清除模板中的演示值，但保留工作表和工作簿级属性。
        for row in worksheet.iter_rows():
            for cell in row:
                cell.value = None

        headers = ["UAV ID"] + [
            f"{index}th Inspection Point" for index in range(1, max_stops + 1)
        ]
        for column, value in enumerate(headers, start=1):
            cell = worksheet.cell(1, column, value)
            cell._style = copy(template_style)

        for row_index, (route, visits) in enumerate(
            zip(routes, visits_by_route), start=2
        ):
            values = [int(route["uav"]), *visits]
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row_index, column, value)
                cell._style = copy(template_style)

        # 原始示例模板有“…”占位行；当某算例只有两架无人机时，必须显式删除该空行，
        # 否则 Excel 的 used range 仍会把它视作正式结果的一部分。
        expected_last_row = len(routes) + 1
        if worksheet.max_row > expected_last_row:
            worksheet.delete_rows(
                expected_last_row + 1,
                worksheet.max_row - expected_last_row,
            )

        worksheet.column_dimensions["A"].width = template_width_a
        for column in range(2, max_stops + 2):
            worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = (
                template_width_route
            )
        # 官方空模板未设置冻结窗格或筛选器，正式文件保持同样的界面属性。
        worksheet.freeze_panes = None
        worksheet.auto_filter.ref = None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify_formal_workbook(
    workbook_path: Path,
    results: dict,
    problem: int,
) -> None:
    """逐格回读正式工作簿，确认没有额外汇总列且路线与 JSON 一致。"""
    workbook = load_workbook(workbook_path, data_only=False)
    if tuple(workbook.sheetnames) != CASE_NAMES:
        raise AssertionError(f"{workbook_path.name}: 工作表必须为 {CASE_NAMES}")

    for case_name in CASE_NAMES:
        worksheet = workbook[case_name]
        case = _case_payload(results, case_name, problem)
        routes = sorted(case["routes"], key=lambda item: int(item["uav"]))
        expected_visits = [_inspection_sequence(route) for route in routes]
        max_stops = max(len(visits) for visits in expected_visits)
        expected_headers = ["UAV ID"] + [
            f"{index}th Inspection Point" for index in range(1, max_stops + 1)
        ]
        actual_headers = [
            worksheet.cell(1, column).value for column in range(1, max_stops + 2)
        ]
        if actual_headers != expected_headers:
            raise AssertionError(f"{workbook_path.name}/{case_name}: 表头不符合模板")
        if worksheet.max_column != max_stops + 1:
            raise AssertionError(f"{workbook_path.name}/{case_name}: 存在额外结果列")
        if worksheet.max_row != len(routes) + 1:
            raise AssertionError(f"{workbook_path.name}/{case_name}: UAV 行数不一致")

        for row_index, (route, visits) in enumerate(
            zip(routes, expected_visits), start=2
        ):
            actual = [
                worksheet.cell(row_index, column).value
                for column in range(1, max_stops + 2)
            ]
            expected = [int(route["uav"]), *visits] + [None] * (
                max_stops - len(visits)
            )
            if actual != expected:
                raise AssertionError(
                    f"{workbook_path.name}/{case_name}/UAV{route['uav']}: 路线不一致"
                )
